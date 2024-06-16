from openpyxl import load_workbook
import ezdxf
from ezdxf.enums import TextEntityAlignment
file_path = '/Users/shubham/Downloads/Music_player/pyautocad/geological_log.xlsx'
workbook = load_workbook(filename=file_path, data_only=True)
print("Sheet names:", workbook.sheetnames)
sheets = workbook.sheetnames
collar_data = []
sheet = workbook[sheets[0]]
for row in sheet.iter_rows(values_only=True):
    BHID = row[1]
    XCOLLAR = row[4]
    YCOLLAR = row[5]
    ZCOLLAR = row[6]
    DEPTH = row[7]
    collar_row = (BHID, XCOLLAR, YCOLLAR, ZCOLLAR, DEPTH)
    collar_data.append(collar_row)
# print("Collar Data:")
# for row in collar_data:
#     print(row)
litho_data = []
sheet = workbook[sheets[1]]
temp=""
temp_arr=[]
for row in sheet.iter_rows(values_only=True):
    BHID = row[0]
    if BHID!=temp:
        temp=BHID
        litho_data.append(temp_arr)
        temp_arr=[]
    FROM = row[1]
    TO = row[2]
    DEPTH = row[3]
    DESC = row[4]
    LITHOLOGY = row[5]
    LITHOID = row[6]
    litho_row = (BHID, FROM, TO, DEPTH, DESC, LITHOLOGY, LITHOID)
    temp_arr.append(litho_row)
# litho_data.append(litho_row)
# print("Litho Data:")
# for row in litho_data:
#     print(row)
#     # for i in row:
#     #     print(i)
#     #     print("\n")
#     print("\n");
## Starting the DXF file creation

# print(litho_data[2])
drawings=[]
drawing=litho_data[2];
doc = ezdxf.new(dxfversion='R2010')
msp = doc.modelspace()
x_start = 0
y_start = 0
width = 25

# print(litho_data)
litho_data.pop(0)
litho_data.pop(0)
drawings=litho_data
doc = ezdxf.new(dxfversion='R2010')
msp = doc.modelspace()
x_start = 0
y_start = 0
width = 25
lithology_colors = {
    'TS': 4,            # Cyan
    'OB': 2, # Yellow
    'WM': 3,            # Green
    'SEAM-IX': 1,    # Red
    'SEAM-XI': 1,    # Red
    'SEAM-X': 1,    # Red
    'SEAM-VIII': 1,    # Red
    'IB': 5, # Magenta
    'SST': 6, # Cyan
}
spacing=0
for drawing in drawings:
    text_position = (10+spacing,10)
    text = msp.add_text(drawing[0][0], dxfattribs={'height': 1.5, 'layer': drawing[0][0], 'color': 0})
    text.set_placement(text_position, align=TextEntityAlignment.MIDDLE_CENTER)
    for x in drawing:
        start_depth = float(x[1])
        end_depth = float(x[2])
        thickness = float(x[3])
        lithology = x[5]
        layer=x[6]
        color = lithology_colors.get(lithology, 1)
        p1=(x_start+spacing, y_start - start_depth)
        p2=(x_start+spacing + width, y_start - start_depth)
        p3=(x_start+spacing + width, y_start - start_depth - thickness)
        p4=(x_start+spacing, y_start - start_depth - thickness)
        try:
            msp.add_lwpolyline([p1, p2, p3, p4, p1], dxfattribs={'layer': 'Layer1', 'color': color})
            hatch = msp.add_hatch(color=color)
            hatch.paths.add_polyline_path([p1, p2, p3, p4, p1])
            # add text to lithology
            text_position = (x_start +spacing+ width + 2, y_start - (start_depth + thickness / 2))
            text = msp.add_text(lithology, dxfattribs={'height': 1.5, 'layer': layer, 'color': 0})
            text.set_placement(text_position, align=TextEntityAlignment.LEFT)
        except:
            print("Error")
    spacing+=100
doc.saveas('geological_log.dxf')
print("DXF file 'geological_log.dxf' created successfully.")
# print(litho_data[0])
